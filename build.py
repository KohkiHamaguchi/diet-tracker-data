#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
diet.json を正本として、
  - data/diet.xlsx   … 人が開く用のアーカイブ（openpyxl）
  - index.html       … ブラウザで見れる自己完結ビューア（外部依存なし）
を生成する。GitHub Actions から push 検知で実行される。
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data" / "diet.json"
XLSX = ROOT / "data" / "diet.xlsx"
HTML = ROOT / "index.html"

CAL_TARGET = 1800
BMR = 1800
KCAL_PER_STEP = 0.04

EX_LABEL = {
    "squat": "スクワット", "push": "プッシュアップ", "plank": "プランク",
    "lunge": "ランジ", "hip": "ヒップリフト", "crunch": "クランチ",
    "jog": "ジョギング", "stair": "階段昇降", "rope": "縄跳び", "stretch": "ストレッチ",
}
EX_KCAL = {
    "squat": 50, "push": 40, "plank": 20, "lunge": 45, "hip": 30,
    "crunch": 35, "jog": 150, "stair": 90, "rope": 130, "stretch": 30,
}


def load():
    if DATA.exists():
        try:
            return json.loads(DATA.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"goal": 69, "entries": {}}


# ── Excel ──────────────────────────────────────────────
def build_xlsx(d):
    wb = Workbook()
    ws = wb.active
    ws.title = "diet"

    headers = ["日付", "体重(kg)", "摂取kcal", "歩数", "消費kcal", "運動", "食事内訳"]
    ws.append(headers)
    thin = Side(style="thin", color="13263A")
    for c in ws[1]:
        c.font = Font(bold=True, color="E4F0F8")
        c.fill = PatternFill("solid", fgColor="0C1824")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=thin)

    for date in sorted(d.get("entries", {})):
        e = d["entries"][date]
        types = e.get("types", [])
        burn = BMR + round((e.get("s") or 0) * KCAL_PER_STEP) + sum(EX_KCAL.get(t, 0) for t in types)
        types_str = "・".join(EX_LABEL.get(t, t) for t in types)
        items_str = ", ".join(f'{i.get("name","")}({i.get("kcal",0)})' for i in e.get("items", []))
        ws.append([
            date, e.get("w", ""), e.get("c", ""), e.get("s", ""),
            burn if types or e.get("s") else "", types_str, items_str,
        ])

    for i, w in enumerate([12, 10, 10, 8, 10, 22, 44], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)


# ── HTML（自己完結・SVGスパークラインはPythonで生成）──────────
def sparkline(weights, goal, w=620, h=160, pad=16):
    """weights: [(date, value), ...] 昇順。SVGの折れ線を返す。"""
    if len(weights) < 2:
        return '<div class="empty">体重の記録が2件以上になると推移が表示されます</div>'
    vals = [v for _, v in weights]
    lo, hi = min(min(vals), goal), max(max(vals), goal)
    if hi - lo < 1:
        hi += 1
        lo -= 1
    n = len(weights)

    def x(i):
        return pad + (w - 2 * pad) * (i / (n - 1))

    def y(v):
        return pad + (h - 2 * pad) * (1 - (v - lo) / (hi - lo))

    pts = " ".join(f"{x(i):.1f},{y(v):.1f}" for i, (_, v) in enumerate(weights))
    gy = y(goal)
    dots = "".join(f'<circle cx="{x(i):.1f}" cy="{y(v):.1f}" r="2.5" fill="#4ecca3"/>'
                   for i, (_, v) in enumerate(weights)) if n <= 60 else ""
    return f'''<svg viewBox="0 0 {w} {h}" class="spark" preserveAspectRatio="none">
      <line x1="{pad}" y1="{gy:.1f}" x2="{w-pad}" y2="{gy:.1f}" stroke="#ffb347" stroke-dasharray="5 5" stroke-width="1"/>
      <polyline points="{pts}" fill="none" stroke="#4ecca3" stroke-width="2.5"/>
      {dots}
    </svg>
    <div class="legend"><span style="color:#4ecca3">— 体重</span><span style="color:#ffb347">- - 目標 {goal}kg</span></div>'''


def build_html(d):
    entries = d.get("entries", {})
    goal = d.get("goal", 69)
    dates = sorted(entries)
    weighed = [(dt, entries[dt]["w"]) for dt in dates if isinstance(entries[dt].get("w"), (int, float))]
    latest = weighed[-1][1] if weighed else None
    start = weighed[0][1] if weighed else None
    remain = (latest - goal) if latest is not None else None

    rows = []
    for dt in reversed(dates[-60:]):  # 直近60件を新しい順
        e = entries[dt]
        types = "・".join(EX_LABEL.get(t, t) for t in e.get("types", []))
        items = ", ".join(f'{i.get("name","")}<span class="k">{i.get("kcal",0)}</span>' for i in e.get("items", []))
        rows.append(
            f'<tr><td>{dt}</td>'
            f'<td class="num">{e.get("w","") if e.get("w") is not None else ""}</td>'
            f'<td class="num">{e.get("c","") if e.get("c") is not None else ""}</td>'
            f'<td class="num">{("{:,}".format(e["s"])) if isinstance(e.get("s"),(int,float)) else ""}</td>'
            f'<td>{types}</td><td class="food">{items}</td></tr>'
        )
    rows_html = "\n".join(rows) if rows else '<tr><td colspan="6" class="empty">まだ記録がありません</td></tr>'

    head = f"{latest:.1f}" if latest is not None else "—"
    remain_html = f"{max(0, remain):.1f}" if remain is not None else "—"
    spark = sparkline(weighed[-90:], goal)

    html = """<!DOCTYPE html>
<html lang="ja"><head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>Diet Tracker</title>
<style>
  :root{ --bg:#07101a; --card:#0c1824; --border:#13263a; --accent:#4ecca3; --warn:#ffb347; --text:#e4f0f8; --mid:#5a8499; }
  *{box-sizing:border-box;margin:0;padding:0}
  body{background:var(--bg);color:var(--text);font-family:'DM Mono',ui-monospace,monospace;max-width:680px;margin:0 auto;padding:16px}
  .label{font-size:10px;letter-spacing:3px;color:var(--mid);font-weight:600}
  .card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:16px;margin-bottom:12px}
  .big{font-size:44px;font-weight:800;line-height:1}
  .unit{font-size:14px;color:var(--mid);margin-left:4px}
  .row{display:flex;justify-content:space-between;align-items:flex-end}
  .spark{width:100%;height:160px;display:block;margin-top:12px}
  .legend{display:flex;gap:16px;margin-top:6px;font-size:10px}
  table{width:100%;border-collapse:collapse;margin-top:12px;font-size:12px}
  th{color:var(--mid);font-weight:600;text-align:left;padding:8px 6px;border-bottom:1px solid var(--border);position:sticky;top:0;background:var(--card)}
  td{padding:8px 6px;border-bottom:1px solid var(--border);color:var(--text);vertical-align:top}
  td.num{text-align:right;color:var(--accent)}
  td.food{color:var(--mid)}
  td.food .k{color:var(--mid);font-size:10px;margin-left:4px}
  .empty{color:var(--mid);text-align:center;padding:20px 0;font-size:12px}
  .foot{color:#1e3448;font-size:10px;text-align:center;margin-top:8px}
</style></head>
<body>
  <div class="card">
    <div class="label">DIET TRACKER</div>
    <div class="row" style="margin-top:8px">
      <div><span class="big">__HEAD__</span><span class="unit">kg</span></div>
      <div style="text-align:right"><span class="big" style="color:var(--warn);font-size:30px">__REMAIN__</span><span class="unit">kg 残り</span></div>
    </div>
  </div>
  <div class="card">
    <div class="label">体重の推移</div>
    __SPARK__
  </div>
  <div class="card">
    <div class="label">記録一覧（直近）</div>
    <div style="overflow-x:auto"><table>
      <thead><tr><th>日付</th><th>体重</th><th>kcal</th><th>歩数</th><th>運動</th><th>食事</th></tr></thead>
      <tbody>
      __ROWS__
      </tbody>
    </table></div>
  </div>
  <div class="foot">自動生成 / 正本は data/diet.json</div>
</body></html>"""

    html = (html
            .replace("__HEAD__", head)
            .replace("__REMAIN__", remain_html)
            .replace("__SPARK__", spark)
            .replace("__ROWS__", rows_html))
    HTML.write_text(html, encoding="utf-8")


def main():
    d = load()
    build_xlsx(d)
    build_html(d)
    print(f"built: {len(d.get('entries', {}))} entries -> diet.xlsx + index.html")


if __name__ == "__main__":
    main()
